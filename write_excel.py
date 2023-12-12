import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import os
from write2 import main

def apply_border_to_merged_cells(sheet, start_row, end_row, start_col, end_col):
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border

# Define the absolute file paths
excel_file = os.path.abspath('test.xlsx')
pdf_file = os.path.abspath('test .pdf')

# Check if the Excel file exists
if not os.path.exists(excel_file):
    # Create a new workbook if it doesn't exist
    wb = Workbook()
    sheet = wb.active

    # Merge cells B8 to I8 and J8 to L8
    sheet.merge_cells('B8:I8')
    sheet.merge_cells('J8:L8')
    sheet.merge_cells('M8:P8')
    sheet.merge_cells('F10:G10')

    # Set the row height to make sure both merged cells are visible
    sheet.row_dimensions[8].height = 30

    # Apply borders to merged cells
    apply_border_to_merged_cells(sheet, 8, 8, 2, 9)  # Apply borders to B8:I8
    apply_border_to_merged_cells(sheet, 8, 8, 10, 12)  # Apply borders to J8:L8
    apply_border_to_merged_cells(sheet, 8, 8, 13, 16)  # Apply borders to M8:P8
    apply_border_to_merged_cells(sheet, 10, 10, 2, 3)  # Apply borders to M8:P
    apply_border_to_merged_cells(sheet, 10, 10, 4, 5)  # Apply borders to M8:P
    apply_border_to_merged_cells(sheet, 10, 10, 6, 8)  # Apply borders to M8:P

    # Set the values in the merged cells
    sheet['B8'] = 'GTI : 8.28.165 FUEL CALIBRATION.test'
    sheet['J8'] = 'UNIT'
    sheet['M8'] = 'Engine Type'
    sheet['B10'] = 'A/C'
    sheet['D10'] = 'VERSION'
    sheet['F10'] = 'FQI P/N'
    
    

    # Save the workbook
    wb.save(excel_file)

else:
    # Open the existing workbook
    wb = xw.Book(excel_file)

    # Define the sheet after opening the workbook
    sheet = wb.sheets[0]

    # Update the values in the merged cells
    sheet.range('B8').value = 'GTI : 8.28.165 FUEL CALIBRATION.test2'
    sheet.range('J8').value = 'UNIT2'
    sheet.range('M8').value = 'Engine Type'

    # Apply borders again after updating values
    apply_border_to_merged_cells(sheet, 8, 8, 2, 9)
    apply_border_to_merged_cells(sheet, 8, 8, 10, 12)
    apply_border_to_merged_cells(sheet, 8, 8, 13, 16)  # Apply borders to M8:P8

    # Save the workbook
    wb.save(excel_file)
    wb.close()

# Create PDF from Excel using built-in functionality
app = xw.App(visible=True)
wb = xw.Book(excel_file)

# Activate the appropriate sheet before exporting
sheet = wb.sheets[0]
sheet.activate()

os.startfile(excel_file)
main()
